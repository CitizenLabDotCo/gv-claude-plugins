"""
Spreadsheet generator for the Go Vocal competitor-battlecard skill.

Reads a briefing JSON, produces a branded .xlsx with one Overview tab and
six canonical product tabs. Tab names, column headers, legend and category
names are localised via scripts/localisation.py. Row content (requirement
names, status emoji, details) comes straight from the briefing — the
briefing author writes that in the target language.

Usage
-----
    python generate_spreadsheet.py <briefing.json> <output_path.xlsx>

Briefing schema (tabs section)
------------------------------
{
  "competitor": "Decidim",
  "language": "en",
  "tabs": {
    "1": {
      "categories": [
        {
          "name": "IDEATION / IDEAS BOX",
          "rows": [
            {
              "requirement": "Citizens can submit ideas...",
              "gv_status": "🟢", "gv_detail": "Full support...",
              "competitor_status": "🟢", "competitor_detail": "Proposals component...",
              "source": "https://docs.decidim.org/proposals"
            },
            ...
          ]
        }
      ]
    },
    "2": { ... },  "3": { ... },  "4": { ... },  "5": { ... },  "6": { ... }
  }
}

Tabs with no categories are skipped — generating a battlecard for a
competitor that doesn't have, say, a back-office (tab 2) is allowed.

Status emoji are the exact unicode characters 🟢 🟡 🔴 — the generator
looks them up to pick the cell fill (green / yellow / red).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Make the localisation module importable whether we're launched from
# scripts/ or from the skill root.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from localisation import L  # noqa: E402


# ---------------------------------------------------------------------------
# Brand tokens
# ---------------------------------------------------------------------------
DARK_PURPLE = "1E155D"
MEDIUM_PURPLE = "43369B"
CHERRY = "FF3E52"
LILAC_1 = "F0EEFA"
WHITE = "FFFFFF"
ALT = "FAFAFA"
GREEN = "C8E6C9"
YELLOW = "FFF3B0"
RED = "F5C6CB"

HEADER_FILL = PatternFill("solid", fgColor=DARK_PURPLE)
CATEGORY_FILL = PatternFill("solid", fgColor=LILAC_1)
ALT_FILL = PatternFill("solid", fgColor=ALT)
GREEN_FILL = PatternFill("solid", fgColor=GREEN)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW)
RED_FILL = PatternFill("solid", fgColor=RED)

HEADER_FONT = Font(name="Chivo", bold=True, color=WHITE, size=10)
CATEGORY_FONT = Font(name="Chivo", bold=True, color=DARK_PURPLE, size=10)
REQ_FONT = Font(name="Chivo", size=9, color=DARK_PURPLE)
BOLD_FONT = Font(name="Chivo", bold=True, size=9, color=DARK_PURPLE)
STATUS_FONT = Font(name="Chivo", size=14, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="DCDCDC"),
    right=Side(style="thin", color="DCDCDC"),
    top=Side(style="thin", color="DCDCDC"),
    bottom=Side(style="thin", color="DCDCDC"),
)

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Status helpers
# ---------------------------------------------------------------------------
GREEN_EMOJI = "\U0001F7E2"
YELLOW_EMOJI = "\U0001F7E1"
RED_EMOJI = "\U0001F534"


def normalise_status(s: str) -> str:
    """Accept several common encodings for the three statuses."""
    if s is None:
        return YELLOW_EMOJI
    s = s.strip()
    if s in {GREEN_EMOJI, "🟢", "green", "GREEN", "ok", "OK"}:
        return GREEN_EMOJI
    if s in {YELLOW_EMOJI, "🟡", "yellow", "YELLOW", "partial", "unverified"}:
        return YELLOW_EMOJI
    if s in {RED_EMOJI, "🔴", "red", "RED", "no", "NO"}:
        return RED_EMOJI
    return YELLOW_EMOJI


def fill_for_status(status: str) -> PatternFill:
    if status == GREEN_EMOJI:
        return GREEN_FILL
    if status == YELLOW_EMOJI:
        return YELLOW_FILL
    return RED_FILL


# ---------------------------------------------------------------------------
# Sheet setup
# ---------------------------------------------------------------------------
def setup_sheet(ws, competitor: str, lang: str) -> int:
    """Configure columns and write the header row. Returns next row to write."""
    ws.column_dimensions["A"].width = 6   # #
    ws.column_dimensions["B"].width = 40  # Requirement
    ws.column_dimensions["C"].width = 8   # GV
    ws.column_dimensions["D"].width = 50  # Go Vocal — Details
    ws.column_dimensions["E"].width = 8   # Competitor
    ws.column_dimensions["F"].width = 50  # Competitor — Details

    headers = [
        L(lang, "col_num"),
        L(lang, "col_requirement"),
        L(lang, "col_gv"),
        L(lang, "col_gv_detail"),
        competitor,
        L(lang, "col_comp_detail_fmt", competitor=competitor),
    ]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    return 2


def add_category(ws, row: int, text: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = CATEGORY_FONT
    cell.fill = CATEGORY_FILL
    cell.alignment = Alignment(vertical="center", indent=1)
    cell.border = THIN_BORDER
    for col in range(2, 7):
        ws.cell(row=row, column=col).fill = CATEGORY_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.row_dimensions[row].height = 24
    return row + 1


def add_requirement_row(
    ws,
    row: int,
    num: int,
    requirement: str,
    gv_status: str,
    gv_detail: str,
    comp_status: str,
    comp_detail: str,
) -> int:
    ws.cell(row=row, column=1, value=num).font = BOLD_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN

    ws.cell(row=row, column=2, value=requirement).font = REQ_FONT
    ws.cell(row=row, column=2).alignment = WRAP_ALIGN

    gv_cell = ws.cell(row=row, column=3, value=gv_status)
    gv_cell.font = STATUS_FONT
    gv_cell.alignment = CENTER_ALIGN
    gv_cell.fill = fill_for_status(gv_status)

    ws.cell(row=row, column=4, value=gv_detail).font = REQ_FONT
    ws.cell(row=row, column=4).alignment = WRAP_ALIGN

    comp_cell = ws.cell(row=row, column=5, value=comp_status)
    comp_cell.font = STATUS_FONT
    comp_cell.alignment = CENTER_ALIGN
    comp_cell.fill = fill_for_status(comp_status)

    ws.cell(row=row, column=6, value=comp_detail).font = REQ_FONT
    ws.cell(row=row, column=6).alignment = WRAP_ALIGN

    for col in range(1, 7):
        ws.cell(row=row, column=col).border = THIN_BORDER

    # Alternating row shade (skips the status columns — their fill is
    # already driven by the status emoji).
    if num % 2 == 0:
        for col in (1, 2, 4, 6):
            ws.cell(row=row, column=col).fill = ALT_FILL

    ws.row_dimensions[row].height = 66
    return row + 1


# ---------------------------------------------------------------------------
# Overview tab
# ---------------------------------------------------------------------------
def build_overview(wb, briefing: dict, lang: str) -> None:
    competitor = briefing.get("competitor", "")
    ws = wb.active
    ws.title = L(lang, "tab_overview")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 84

    # Title
    ws.merge_cells("B2:C2")
    t = ws["B2"]
    t.value = L(lang, "sheet_title", competitor=competitor)
    t.font = Font(name="Chivo", size=20, color=DARK_PURPLE)

    ws.merge_cells("B3:C3")
    st = ws["B3"]
    st.value = L(lang, "sheet_subtitle", competitor=competitor)
    st.font = Font(name="Chivo", size=11, color=MEDIUM_PURPLE, italic=True)
    st.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 36

    # Legend
    ws["B5"] = L(lang, "legend_title")
    ws["B5"].font = Font(name="Chivo", size=13, bold=True, color=DARK_PURPLE)

    legend_rows = [
        (GREEN_EMOJI, L(lang, "legend_green")),
        (YELLOW_EMOJI, L(lang, "legend_yellow")),
        (RED_EMOJI, L(lang, "legend_red")),
    ]
    r = 6
    for icon, desc in legend_rows:
        ws.cell(row=r, column=2, value=icon).font = Font(name="Chivo", size=16)
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=r, column=3, value=desc).font = Font(
            name="Chivo", size=11, color=DARK_PURPLE
        )
        ws.cell(row=r, column=3).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 22
        r += 1

    # How to read this document — the six-tab map
    r += 1
    ws.cell(row=r, column=2, value=L(lang, "how_to_read")).font = Font(
        name="Chivo", size=13, bold=True, color=DARK_PURPLE
    )
    r += 1
    tab_map = [
        ("tab_1", "tab_1_desc"),
        ("tab_2", "tab_2_desc"),
        ("tab_3", "tab_3_desc"),
        ("tab_4", "tab_4_desc"),
        ("tab_5", "tab_5_desc"),
        ("tab_6", "tab_6_desc"),
    ]
    for tab_key, desc_key in tab_map:
        r += 1
        ws.cell(row=r, column=2, value=L(lang, tab_key)).font = Font(
            name="Chivo", size=10, bold=True, color=DARK_PURPLE
        )
        ws.cell(row=r, column=3, value=L(lang, desc_key)).font = Font(
            name="Chivo", size=10, color=MEDIUM_PURPLE
        )

    # Honesty note (cherry italic)
    r += 3
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    note = ws.cell(row=r, column=2)
    note.value = L(lang, "honesty_note_fmt", competitor=competitor)
    note.font = Font(name="Chivo", size=10, italic=True, color=CHERRY)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30


# ---------------------------------------------------------------------------
# Product tabs
# ---------------------------------------------------------------------------
def build_product_tab(
    wb, tab_number: int, tab_data: dict, briefing: dict, lang: str
) -> None:
    categories = tab_data.get("categories", []) if tab_data else []
    if not categories:
        return  # nothing to say about this product area for this competitor

    competitor = briefing.get("competitor", "")
    ws = wb.create_sheet(L(lang, f"tab_{tab_number}"))
    row = setup_sheet(ws, competitor, lang)

    n = 1
    for category in categories:
        row = add_category(ws, row, category.get("name", ""))
        for entry in category.get("rows", []):
            gv_status = normalise_status(entry.get("gv_status", ""))
            comp_status = normalise_status(entry.get("competitor_status", ""))
            row = add_requirement_row(
                ws,
                row,
                n,
                entry.get("requirement", ""),
                gv_status,
                entry.get("gv_detail", ""),
                comp_status,
                entry.get("competitor_detail", ""),
            )
            n += 1


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def build(briefing: dict, output_path: Path) -> None:
    lang = briefing.get("language", "en")
    wb = openpyxl.Workbook()
    build_overview(wb, briefing, lang)

    tabs = briefing.get("tabs", {})
    for i in range(1, 7):
        build_product_tab(wb, i, tabs.get(str(i), {}), briefing, lang)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        sys.stderr.write(
            "Usage: python generate_spreadsheet.py <briefing.json> <output_path.xlsx>\n"
        )
        return 2

    briefing_path = Path(argv[1])
    output_path = Path(argv[2])

    with briefing_path.open("r", encoding="utf-8") as f:
        briefing = json.load(f)

    build(briefing, output_path)
    print(f"Saved: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
